import json
import os
import hashlib
from openpyxl import load_workbook

INPUT_XLSX = "単語リスト.xlsx"
OUT_DIR = "data"

def book_id_from_title(title: str) -> str:
    # シート名が日本語でも安全に扱えるよう、ハッシュからIDを作る
    h = hashlib.sha1(title.encode("utf-8")).hexdigest()[:10]
    return f"book_{h}"

def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    wb = load_workbook(INPUT_XLSX, data_only=True)

    books_index = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # A:番号 B:英単語 C:日本語 という前提で読み取り
        items = []
        row = 1
        while True:
            no = ws.cell(row=row, column=1).value
            en = ws.cell(row=row, column=2).value
            ja = ws.cell(row=row, column=3).value

            # 途中で空行に当たったら終了（A/B/Cが全部空なら終わり）
            if no is None and en is None and ja is None:
                break

            # 多少の空欄はスキップ（念のため）
            if no is None or en is None or ja is None:
                row += 1
                continue

            try:
                no_int = int(no)
            except:
                row += 1
                continue

            items.append({
                "no": no_int,
                "en": str(en).strip(),
                "ja": str(ja).strip()
            })
            row += 1

        if not items:
            continue

        # No順にソート（念のため）
        items.sort(key=lambda x: x["no"])

        bid = book_id_from_title(sheet_name)
        book_path = os.path.join(OUT_DIR, f"{bid}.json")

        with open(book_path, "w", encoding="utf-8") as f:
            json.dump(
                {"title": sheet_name, "items": items},
                f,
                ensure_ascii=False,
                indent=2
            )

        books_index.append({
            "id": bid,
            "title": sheet_name,
            "count": len(items)
        })

    # 単語帳一覧（プルダウン用）
    books_index.sort(key=lambda x: x["title"])
    with open(os.path.join(OUT_DIR, "books.json"), "w", encoding="utf-8") as f:
        json.dump(books_index, f, ensure_ascii=False, indent=2)

    print("OK: data/books.json と各 book_XXXX.json を生成しました。")
    print(f"生成した単語帳数: {len(books_index)}")

if __name__ == "__main__":
    main()
